#!/usr/bin/env python3
"""
extract_qualitivity.py — Extract data from Qualitivity .xlsm and generate JS data block

Usage:
    python extract_qualitivity.py Qualitivity_0306.xlsm > data_block.js

    Then paste the output between <script id="ds"> and </script> in the HTML file.

Requires: pip install openpyxl
"""

import openpyxl
import json
import sys
from datetime import datetime, date
from collections import Counter


def extract_records(wb, sheet_name, include_comment=True, max_cols=93, dc_filter=False):
    """Extract claim records from a Qualitivity RO sheet.

    Args:
        dc_filter: If True, read column 96 (Status) and only include rows where Status='Include'.
                   Used for RO (DC) sheet to match official Qualitivity report numbers.
    """
    key_map = {
        0: 'no', 1: 'brand', 5: 'proj', 4: 'model',
        10: 'partNo', 11: 'keyParts', 12: 'safety',
        13: 'partNM', 14: 'system', 15: 'natCode', 16: 'natName',
        20: 'causeCode', 22: 'comment',
        35: 'confMonth',
        36: 'stockP', 37: 'useP', 38: 'mileage',
        39: 'claims', 40: 'partCost', 41: 'laborCost', 42: 'outsource', 43: 'totalCost',
        46: 'nation', 47: 'region', 50: 'dealer',
        57: 'devName', 60: 'faultCorp', 65: 'vin'
    }

    # For DC sheets, we need to read up to column 97 (index 96 = Status)
    read_cols = max(max_cols, 97) if dc_filter else max_cols

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    records = []
    skipped = 0

    for row in rows[1:]:
        vals = list(row)[:read_cols]
        if vals[0] is None:
            continue

        # DC Status filter: column 96 must be 'Include'
        if dc_filter:
            status = vals[96] if len(vals) > 96 else None
            if status != 'Include':
                skipped += 1
                continue

        r = {}
        for idx, key in key_map.items():
            v = vals[idx] if idx < len(vals) else None
            if isinstance(v, (datetime, date)):
                v = v.strftime('%Y-%m-%d')
            if key == 'comment':
                if not include_comment:
                    continue
                if v and isinstance(v, str):
                    v = v[:120]
            if v is None:
                continue
            r[key] = v
        # Derive state from dealer code
        dealer = str(r.get('dealer', ''))
        if len(dealer) >= 2 and dealer[:2].isalpha():
            r['state'] = dealer[:2].upper()
        records.append(r)

    if dc_filter and skipped > 0:
        print(f"//   DC filter: {skipped} 'Not Include' records skipped", file=sys.stderr)

    return records


def extract_sales_by_state(wb):
    """Extract US vehicle sales aggregated by state from DB Sales sheet."""
    ws = wb['DB Sales (DC)']
    rows = list(ws.iter_rows(values_only=True))
    sales = Counter()

    for row in rows[1:]:
        vals = list(row)[:30]
        if vals[0] is None:
            continue
        nation = vals[24]  # Sale nations column
        dealer = str(vals[20] or '')
        if nation == 'United States' and len(dealer) >= 2 and dealer[:2].isalpha():
            sales[dealer[:2].upper()] += 1

    return dict(sales)


def main():
    if len(sys.argv) < 2:
        print("Usage: python extract_qualitivity.py <path_to_xlsm>", file=sys.stderr)
        sys.exit(1)

    filepath = sys.argv[1]
    print(f"// Data extracted from {filepath}", file=sys.stderr)
    print(f"// Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", file=sys.stderr)

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    # Extract claim records
    # DC: filter by Status column (col 96) = 'Include' to match official Qualitivity numbers
    d3m = extract_records(wb, 'RO (3M)', include_comment=True)
    ddc = extract_records(wb, 'RO (DC)', include_comment=True, dc_filter=True)
    d12m = extract_records(wb, 'RO (12M)', include_comment=False)

    print(f"// 3M: {len(d3m)} records", file=sys.stderr)
    print(f"// DC: {len(ddc)} records (filtered by Status='Include')", file=sys.stderr)
    print(f"// 12M: {len(d12m)} records", file=sys.stderr)

    # Build compact 12M format
    cols_12m = ['proj', 'partNM', 'system', 'natCode', 'natName', 'causeCode', 'confMonth',
                'useP', 'mileage', 'claims', 'totalCost', 'nation', 'region', 'dealer', 'devName', 'state']
    rows_12m = [[r.get(c, '') for c in cols_12m] for r in d12m]

    # Extract sales
    sales = extract_sales_by_state(wb)
    print(f"// US Sales: {sum(sales.values())} across {len(sales)} states", file=sys.stderr)

    wb.close()

    # Output JS
    js = f"const QUALITIVITY_DATA = {json.dumps({'3M': d3m, 'DC': ddc}, ensure_ascii=False)};\n"
    js += f"const QUALITIVITY_12M_COLS = {json.dumps(cols_12m)};\n"
    js += f"const QUALITIVITY_12M_ROWS = {json.dumps(rows_12m, ensure_ascii=False)};\n"
    js += f"const US_SALES_BY_STATE = {json.dumps(sales)};\n"

    print(js)

    print(f"// Total JS output: {len(js)/1024/1024:.2f} MB", file=sys.stderr)


if __name__ == '__main__':
    main()
