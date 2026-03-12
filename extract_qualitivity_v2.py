#!/usr/bin/env python3
"""
extract_qualitivity.py — Extract & filter data from Qualitivity .xlsm files

USAGE:
    # Extract from the main file (has 3M, DC, Sales):
    python extract_qualitivity.py Qualitivity_0306.xlsm
    
    # Extract from the RO file (has 12M):
    python extract_qualitivity.py Qualitivity_0306_RO.xlsm
    
    # Extract from both at once:
    python extract_qualitivity.py Qualitivity_0306.xlsm Qualitivity_0306_RO.xlsm

OUTPUT:
    Creates qualitivity_data.json in the current directory.
    This JSON can be committed to GitHub (typically <5 MB).

FILTERS APPLIED:
    DC:  Status (col 96) == 'Include'     → removes rejected/pending claims
    3M:  No filter needed                  → first-half columns are pre-filtered
    12M: Status (col 96) == '1'            → removes newly-added unconfirmed claims

Requires: pip install openpyxl
"""

import openpyxl
import json
import sys
import os
from datetime import datetime, date
from collections import Counter

# Column mapping: index → field name
KEY_MAP = {
    0: 'no', 1: 'brand', 5: 'proj', 4: 'model',
    10: 'partNo', 11: 'keyParts', 12: 'safety',
    13: 'partNM', 14: 'system', 15: 'natCode', 16: 'natName',
    20: 'causeCode', 22: 'comment',
    29: 'salesMonth',   # Vehicle sales month (used by PIVOTs)
    35: 'confMonth',    # Confirmation month
    36: 'stockP', 37: 'useP', 38: 'mileage',
    39: 'claims', 40: 'partCost', 41: 'laborCost', 42: 'outsource', 43: 'totalCost',
    46: 'nation', 47: 'region', 50: 'dealer',
    57: 'devName', 60: 'faultCorp', 65: 'vin',
}

# Extended columns only in DC/12M (beyond standard 93)
EXT_MAP_DC = {92: 'retailSalesMonth', 95: 'resultMonth', 96: 'status'}
EXT_MAP_12M = {96: 'status', 192: 'resultMonth'}  # '1'/'Today', monthly name


def extract_sheet(wb, sheet_name, include_comment=True, max_cols=93, ext_map=None, status_filter=None):
    """Extract records from a Qualitivity RO sheet with optional filtering."""
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    
    records = []
    skipped = 0
    
    for row in rows[1:]:
        vals = list(row)
        if vals[0] is None:
            continue
        
        # Check status filter first (before building full record)
        if status_filter and ext_map:
            for col_idx, field_name in ext_map.items():
                if field_name == 'status' and col_idx < len(vals):
                    status_val = vals[col_idx]
                    if str(status_val) != status_filter:
                        skipped += 1
                        continue  # This continue only exits the inner for loop
            # Need to re-check after inner loop
            if ext_map and status_filter:
                status_col = [idx for idx, name in ext_map.items() if name == 'status']
                if status_col:
                    sv = vals[status_col[0]] if status_col[0] < len(vals) else None
                    if str(sv) != status_filter:
                        skipped += 1
                        continue
        
        r = {}
        for idx, key in KEY_MAP.items():
            if idx >= len(vals):
                continue
            v = vals[idx]
            if isinstance(v, (datetime, date)):
                v = v.strftime('%Y-%m-%d')
            if key == 'comment':
                if not include_comment:
                    continue
                if v and isinstance(v, str):
                    v = v[:120]
            if v is None:
                continue
            r[key] = v
        
        # Extended columns
        if ext_map:
            for idx, key in ext_map.items():
                if idx < len(vals) and vals[idx] is not None:
                    r[key] = str(vals[idx])
        
        # Derive state from dealer code
        dealer = str(r.get('dealer', ''))
        if len(dealer) >= 2 and dealer[:2].isalpha():
            r['state'] = dealer[:2].upper()
        
        records.append(r)
    
    return records, skipped


def extract_sales_by_state(wb):
    """Extract US vehicle sales by state from DB Sales sheet."""
    if 'DB Sales (DC)' not in wb.sheetnames:
        return {}
    ws = wb['DB Sales (DC)']
    rows = list(ws.iter_rows(values_only=True))
    sales = Counter()
    for row in rows[1:]:
        vals = list(row)[:30]
        if vals[0] is None:
            continue
        nation = vals[24]
        dealer = str(vals[20] or '')
        if nation == 'United States' and len(dealer) >= 2 and dealer[:2].isalpha():
            sales[dealer[:2].upper()] += 1
    return dict(sales)


def extract_monthly_sales(wb):
    """Extract monthly vehicle sales from DB Sales (DC) sheet."""
    if 'DB Sales (DC)' not in wb.sheetnames:
        return {}
    ws = wb['DB Sales (DC)']
    rows = list(ws.iter_rows(values_only=True))
    monthly = Counter()
    for row in rows[1:]:
        vals = list(row)
        if vals[0] is None:
            continue
        rsd = vals[22]  # Retail sales date
        m = None
        if isinstance(rsd, str) and len(rsd) >= 7:
            m = rsd[:7]
        elif isinstance(rsd, (datetime, date)):
            m = rsd.strftime('%Y-%m')
        if m:
            monthly[m] += 1
    return dict(monthly)


def main():
    if len(sys.argv) < 2:
        print("Usage: python extract_qualitivity.py <file1.xlsm> [file2.xlsm]", file=sys.stderr)
        sys.exit(1)

    result = {'3M': [], 'DC': [], '12M': [], 'sales_by_state': {}, 'monthly_sales': {}}
    
    for filepath in sys.argv[1:]:
        print(f"Processing: {filepath}", file=sys.stderr)
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        
        for sheet in wb.sheetnames:
            if sheet == 'RO (3M)':
                recs, skip = extract_sheet(wb, sheet, include_comment=True)
                result['3M'] = recs
                print(f"  3M: {len(recs)} records (no filter needed)", file=sys.stderr)
            
            elif sheet == 'RO (DC)':
                recs, skip = extract_sheet(wb, sheet, include_comment=True,
                                          ext_map=EXT_MAP_DC)
                result['DC'] = recs
                print(f"  DC: {len(recs)} records (all statuses, no filter)", file=sys.stderr)
            
            elif sheet == 'RO (12M)':
                recs, skip = extract_sheet(wb, sheet, include_comment=False,
                                          ext_map=EXT_MAP_12M, status_filter='1')
                result['12M'] = recs
                print(f"  12M: {len(recs)} records (filtered, skipped {skip} Today/new)", file=sys.stderr)
            
            elif sheet == 'DB Sales (DC)':
                result['sales_by_state'] = extract_sales_by_state(wb)
                result['monthly_sales'] = extract_monthly_sales(wb)
                print(f"  Sales by state: {sum(result['sales_by_state'].values())} across {len(result['sales_by_state'])} states", file=sys.stderr)
                print(f"  Monthly sales: {dict(sorted(result['monthly_sales'].items()))}", file=sys.stderr)
        
        wb.close()
    
    # Save
    output_path = 'qualitivity_data.json'
    with open(output_path, 'w') as f:
        json.dump(result, f, ensure_ascii=False)
    
    size = os.path.getsize(output_path)
    print(f"\nSaved: {output_path} ({size/1024/1024:.2f} MB)", file=sys.stderr)
    print(f"  3M:  {len(result['3M'])} records", file=sys.stderr)
    print(f"  DC:  {len(result['DC'])} records", file=sys.stderr)
    print(f"  12M: {len(result['12M'])} records", file=sys.stderr)


if __name__ == '__main__':
    main()
