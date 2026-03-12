#!/usr/bin/env python3
"""
inspect_columns.py — Show ALL column headers from Qualitivity .xlsm
Usage: python inspect_columns.py Qualitivity_0306.xlsm
"""
import openpyxl, sys

if len(sys.argv) < 2:
    print("Usage: python inspect_columns.py <path_to_xlsm>")
    sys.exit(1)

wb = openpyxl.load_workbook(sys.argv[1], read_only=True, data_only=True)

for sheet_name in ['RO (3M)', 'RO (DC)', 'RO (12M)']:
    if sheet_name not in wb.sheetnames:
        print(f"\n❌ Sheet '{sheet_name}' not found")
        continue
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True, max_row=3))
    header = rows[0] if rows else []
    sample1 = rows[1] if len(rows) > 1 else []
    sample2 = rows[2] if len(rows) > 2 else []

    print(f"\n{'='*80}")
    print(f"Sheet: {sheet_name} — {len(header)} columns")
    print(f"{'='*80}")

    for i, h in enumerate(header):
        s1 = sample1[i] if i < len(sample1) else ''
        s2 = sample2[i] if i < len(sample2) else ''
        extracted = '✅' if i in {0,1,4,5,10,11,12,13,14,15,16,20,22,35,36,37,38,39,40,41,42,43,46,47,50,57,60,65} else '  '
        # Highlight columns that might be status/decision fields
        flag = ''
        if h and any(kw in str(h).lower() for kw in ['status','decision','valid','approve','reject','void','cancel','accept','result','flag','type','class','categ','judge','confirm']):
            flag = ' ⭐ POSSIBLE STATUS FIELD'
        print(f"  {extracted} [{i:2d}] {str(h)[:50]:<50} | {str(s1)[:30]:<30} | {str(s2)[:30]}{flag}")

# Count rows per sheet
print(f"\n{'='*80}")
print("Row counts:")
for sheet_name in ['RO (3M)', 'RO (DC)', 'RO (12M)']:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        count = sum(1 for row in ws.iter_rows(values_only=True, min_row=2) if row[0] is not None)
        print(f"  {sheet_name}: {count} data rows")

wb.close()
